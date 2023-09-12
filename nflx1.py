import openpyxl


def find_max_english_tv_show_appearances():
    wb = openpyxl.load_workbook("NFLX_DS_data_9_23.xlsx")
    sheet = wb["NFLX Top 10"]

    counts = {}

    for row in sheet.iter_rows():
        if row[0].value == "TV (English)":
            show_title = row[5].value
            if counts.get(show_title) == None:
                counts[show_title] = (1, row[2].value)
            else:
                counts[show_title] = (
                    counts[show_title][0] + 1,
                    counts[show_title][1] + row[2].value,
                )
    wb.close()
    max_ocurances = 0
    max_title = ""

    for key, value in counts.items():
        if value[0] > max_ocurances:
            max_ocurances = value[0]
            max_title = key

    return (max_title, max_ocurances, counts[max_title][1] / max_ocurances)


print("Problem #1:", find_max_english_tv_show_appearances())

class NonEnglishFilm:
    def __init__(self):
        self.h = {}

    def get_imdb_from_title(self, title):
        if self.h == {}:
          wb = openpyxl.load_workbook("NFLX_DS_data_9_23.xlsx")
          sheet = wb["IMDb Rating"]
          for row in sheet.iter_rows():
            self.h[row[0].value] = row[1].value
          wb.close()
        return self.h[title]

    def find_film_non_english_lowest_imbdb(self):
        wb = openpyxl.load_workbook("NFLX_DS_data_9_23.xlsx")
        sheet = wb["NFLX Top 10"]

        films_veiewer_counts_and_imdb = {}

        for row in sheet.iter_rows():
            if row[0].value == "Films (Non-English)":
                film_title = row[5].value
                if films_veiewer_counts_and_imdb.get(film_title) == None:
                    films_veiewer_counts_and_imdb[film_title] = (1, row[2].value)
                else:
                    films_veiewer_counts_and_imdb[film_title] = (
                        films_veiewer_counts_and_imdb[film_title][0] + 1,
                        films_veiewer_counts_and_imdb[film_title][1] + row[2].value,
                    )
        sheet = wb["NFLX Top 10"]
        lowest_imdb = float("inf")
        lowest_imdb_title = ""
        for key in films_veiewer_counts_and_imdb.keys():
            imdb = self.get_imdb_from_title(key)
            if imdb < lowest_imdb:
                lowest_imdb = imdb
                lowest_imdb_title = key
        wb.close()
        return (lowest_imdb_title, lowest_imdb, films_veiewer_counts_and_imdb[lowest_imdb_title][1] / films_veiewer_counts_and_imdb[lowest_imdb_title][0])
              

non_english_film = NonEnglishFilm()
print("Problem #1:", non_english_film.find_film_non_english_lowest_imbdb())

def find_english_film_with_most_weeks_in_top_10():
  wb = openpyxl.load_workbook("NFLX_DS_data_9_23.xlsx")
  sheet = wb["NFLX Top 10"]

  film_weeks_count = {}

  for row in sheet.iter_rows():
    if row[0].value == "Films (English)":
      film_title = row[5].value
      if film_weeks_count.get(film_title) == None:
        film_weeks_count[film_title] = row[1].value
      else:
        film_weeks_count[film_title] += row[1].value

  max_weeks = 0
  max_weeks_title = ""
  for key, value in film_weeks_count.items():
    if value > max_weeks:
      max_weeks = value
      max_weeks_title = key
  
  return (max_weeks_title, max_weeks)

print("Problem #3:", find_english_film_with_most_weeks_in_top_10())